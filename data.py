import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
import json
#define read path
# {path} uses / instead of \
# {path} should start with /Users, not C:\Users
# {path} should end in .xlsx, instead of ending as a folder name
read_path="생산자물가지수.xlsx"

#set workbook
workbook = openpyxl.load_workbook(read_path)

#define sheet name
exisiting_sheet_name="데이터"

#select existing sheet
sheet=workbook[exisiting_sheet_name]

dicts=[]



# #go over each row
# # {index} in the row counter
# # .iter_rows method iterates the row
# # {min_row} is the minimum row where the iteration row starts
# # {start} is the first value of the row counter, {index}
# for col_index, col in enumerate(sheet.iter_cols(min_row=1, values_only=True), start=1):
#     if not col_index%8==2: continue
#     dict={
#         'time': sheet.cell(row=1, column=col_index).value,
#         'pop_15': sheet.cell(row=3, column=col_index).value,
#         'pop_econ': sheet.cell(row=3, column=col_index+1).value,
#         'pop_emp': sheet.cell(row=3, column=col_index+2).value,
#         'pop_unemp': sheet.cell(row=3, column=col_index+3).value,
#         'pop_unecon': sheet.cell(row=3, column=col_index+4).value,
#         'rate_econ': sheet.cell(row=3, column=col_index+5).value,
#         'rate_umemp': sheet.cell(row=3, column=col_index+6).value,
#     }
#     dicts.append(dict)

# json_string = json.dumps(dicts, indent=4)

# with open('data.json', 'w') as f:
#     f.write(json_string)



# with open('data.json', 'r') as f:
#     data = json.load(f)

# for col_index, col in enumerate(sheet.iter_cols(min_col=4, values_only=True), start=4):
#     time=sheet.cell(row=2, column=col_index).value
#     time=time.replace(" 월","")
#     ppi=sheet.cell(row=3, column=col_index).value

#     for item in data:
#         if item.get('time') == time:
#             item['ppi'] = ppi
#             break

# with open('data.json', 'w') as f:
#     json.dump(data, f, indent=4)

# import pandas as pd

# with open('data.json', 'r') as f:
#     data = json.load(f)


# # Read the CSV file
# df = pd.read_csv("traffic.csv")

# # Iterate over rows and access the 'one' column
# for index, row in df.iterrows():
#     times = row['time']
#     times=times.split('-')
#     year=f"20{int(times[1]):02d}"
#     month=f"{int(times[0]):02d}"
#     time=f"{year}.{month}"

#     for item in data:
#         if item.get('time') == time:
#             item['mulga'] = row['mulga']
#             item['inflation'] = row['inflation']
#             break

# with open('data.json', 'w') as f:
#     json.dump(data, f, indent=4)


with open('data.json', 'r') as f:
    data = json.load(f)


    for item in data:
        time=item.get('time')
        item['time'] = time.replace('.','-')

with open('data.json', 'w') as f:
    json.dump(data, f, indent=4)